#!/usr/bin/env python3
"""
创建投递信息Excel模板文件
"""
import os
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("错误: 需要安装 openpyxl 库")
    print("请运行: pip install openpyxl")
    sys.exit(1)

# 获取项目根目录
project_root = Path(__file__).parent.parent.parent
excel_path = project_root / "投递信息管理" / "applications.xlsx"

# 创建工作簿
wb = Workbook()

# 创建主表：applications
ws_applications = wb.active
ws_applications.title = "applications"

# 设置表头
headers = [
    "id",                    # 投递记录ID
    "jd_id",                 # 关联的JD ID（可选）
    "company_name",          # 公司名称
    "position_name",         # 岗位名称
    "application_date",      # 投递日期
    "channel_source",        # 投递渠道/链接
    "status",                # 投递状态
    "work_location",        # 工作地点
    "salary_expectations",   # 薪资待遇
    "next_action_date",     # 下次跟进日期
    "notes",                 # 备注信息
    "created_at",           # 创建时间
    "updated_at"            # 更新时间
]

# 设置表头样式
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

ws_applications.append(headers)
for cell in ws_applications[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# 设置列宽
column_widths = {
    "A": 10,  # id
    "B": 10,  # jd_id
    "C": 20,  # company_name
    "D": 30,  # position_name
    "E": 15,  # application_date
    "F": 40,  # channel_source
    "G": 12,  # status
    "H": 15,  # work_location
    "I": 20,  # salary_expectations
    "J": 15,  # next_action_date
    "K": 50,  # notes
    "L": 20,  # created_at
    "M": 20,  # updated_at
}

for col, width in column_widths.items():
    ws_applications.column_dimensions[col].width = width

# 添加状态说明（在第二行添加注释行，实际使用时可以删除）
status_note = [
    "",
    "",
    "",
    "",
    "",
    "",
    "状态选项: 未投递/已投递/面试中/被拒/拿到Offer",
    "",
    "",
    "",
    "",
    "",
    ""
]
ws_applications.append(status_note)

# 创建时间线表（可选）
ws_timeline = wb.create_sheet("timeline")
timeline_headers = [
    "application_id",  # 关联的投递记录ID
    "date",           # 事件日期
    "event",          # 事件类型
    "notes"           # 事件备注
]
ws_timeline.append(timeline_headers)
for cell in ws_timeline[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# 设置时间线表列宽
ws_timeline.column_dimensions["A"].width = 12
ws_timeline.column_dimensions["B"].width = 15
ws_timeline.column_dimensions["C"].width = 20
ws_timeline.column_dimensions["D"].width = 50

# 创建面试记录表（可选）
ws_interviews = wb.create_sheet("interviews")
interview_headers = [
    "application_id",   # 关联的投递记录ID
    "round",           # 面试轮次
    "date",            # 面试日期
    "type",            # 面试类型
    "interview_notes", # 面经内容
    "questions",       # 面试问题（用分号分隔）
    "feedback",        # 面试反馈
    "result"           # 面试结果
]
ws_interviews.append(interview_headers)
for cell in ws_interviews[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# 设置面试记录表列宽
ws_interviews.column_dimensions["A"].width = 12
ws_interviews.column_dimensions["B"].width = 10
ws_interviews.column_dimensions["C"].width = 15
ws_interviews.column_dimensions["D"].width = 15
ws_interviews.column_dimensions["E"].width = 50
ws_interviews.column_dimensions["F"].width = 50
ws_interviews.column_dimensions["G"].width = 30
ws_interviews.column_dimensions["H"].width = 15

# 保存文件
wb.save(excel_path)
print(f"✅ Excel模板文件已创建: {excel_path}")
print("\n表格说明:")
print("- applications: 主表，记录所有投递信息")
print("- timeline: 时间线表，记录每个投递的时间线事件（可选使用）")
print("- interviews: 面试记录表，记录面试详情（可选使用）")
print("\n💡 提示: 你可以直接在Excel中打开此文件进行编辑")

